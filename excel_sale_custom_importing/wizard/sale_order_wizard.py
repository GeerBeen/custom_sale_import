import openpyxl
import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError

class SaleOrderImportWizard(models.TransientModel):
    _name = "sale.order.import.wizard"
    _description = "Wizard for custom importing Sales from excel."

    excel_file = fields.Binary(string="Excel File", required=True)
    file_name = fields.Char(string="File Name")
    partner_id = fields.Many2one("res.partner", string="Customer", required=True)

    start_row = fields.Integer("Start Row", required=True, default=9)
    product_col = fields.Integer("Product Column", required=True, default=2)
    quantity_col = fields.Integer("Product's Quantity Column", required=True, default=5)
    price_col = fields.Integer("Product's Price Column", required=True, default=6)

    @api.constrains("start_row", "product_col", "quantity_col", "price_col")
    def _check_excel_values(self):
        for record in self:
            if record.start_row < 1 or record.product_col < 1 or record.quantity_col < 1 or record.price_col < 1:
                raise ValidationError(_("Rows and Columns numbers must be bigger than 0!"))

    def _is_positive_float(self, number):
        # якщо не вдається привести, то не число взагалі
        try:
            number = float(number)
        except:
            return False
        if number <= 0:
            return False
        else:
            return True

    def _parse_excel(self) -> dict[str,list]:
        if not self.file_name.endswith(".xlsx"):
            raise ValidationError(_("Only .xlsx files are supported!"))

        # кодований рядок до байтів
        file_bytes = base64.b64decode(self.excel_file)
        # байти до потоку
        file_input = io.BytesIO(file_bytes)
        # поток до об'єкта журналу
        workbook = openpyxl.load_workbook(file_input,data_only=True)
        # перший лист журналу
        sheet = workbook.active


        products = []
        skipped_missing_data_rows = []
        invalid_data_rows = []
        end_row = sheet.max_row

        for current_row in range(self.start_row, end_row + 1):
            # отримуємо поточні значення
            current_product = sheet.cell(row=current_row, column=self.product_col).value
            current_quantity = sheet.cell(row=current_row, column=self.quantity_col).value
            current_price = sheet.cell(row=current_row, column=self.price_col).value

            # # якщо жодного значення нема - виходимо
            # if not current_product and not current_quantity and not current_price:
            #     break

            # якщо всі три є
            if current_product and current_quantity and current_price:
                # якщо ціна або кількість не позитивні дробові, то скіп
                if not self._is_positive_float(current_quantity) or not self._is_positive_float(current_price):
                    invalid_data_rows.append(str(current_row))
                    continue
                # додаємо в список продуктів
                product = (current_product, current_quantity, current_price)
                products.append(product)
            # якщо чогось нема, то зберігаємо для нотаток
            else:
                skipped_missing_data_rows.append(str(current_row))

        return {
            "products" : products,  # рядки з цілими продуктами
            "skipped_missing_data_rows":skipped_missing_data_rows,  # рядки в яких не було деяких значень
            "invalid_data_rows":invalid_data_rows,  # рядки в яких були неправильні значення
        }
    
    def _find_or_create_product(self, product_name, created_products_names):
        product = self.env["product.product"].search([("name","=",product_name)], limit=1)
        if not product:
            product = self.env["product.product"].create({
                "name":product_name,
                "type":"consu"
                })
            created_products_names.append(product_name)
        return product

    def _create_sale_order_line(self, sale_order, product, quantity, price):
            self.env["sale.order.line"].create({
                "order_id": sale_order.id,
                "product_id": product.id,
                "product_uom_qty": float(quantity),
                "price_unit": float(price)
            })

    def _make_notes(self, skipped_missing_data_rows, invalid_data_rows, created_products_names) -> str:
        notes_text = ""
        if skipped_missing_data_rows:
            notes_text += _("Rows that were skipped due to missing data: %s.\n ") % ', '.join(skipped_missing_data_rows)
        if invalid_data_rows:
            notes_text += _("Rows that were skipped due to invalid data: %s.\n ") % ', '.join(invalid_data_rows)
        if created_products_names:
            notes_text += _("Created products: %s.\n ") % ', '.join(created_products_names)
        return notes_text


    def sale_order_import_action(self) -> dict:
        self.ensure_one()

        # отримуємо дані з екселю
        parsed_info = self._parse_excel()
        products_data = parsed_info.get("products", None)
        skipped_missing_data_rows = parsed_info.get("skipped_missing_data_rows", None)
        invalid_data_rows = parsed_info.get("invalid_data_rows", None)
        created_products_names = []

        # якщо продуктів нема, то закінчуємо
        if not products_data:
            raise ValidationError(_("Nothing to import!"))

        # створюємо sale_order об'єкт
        sale_order = self.env['sale.order'].create({
            "partner_id": self.partner_id.id,
        })
        # ітеруємось по даним з екселю, створюємо або знаходимо товари і sale order_lines
        for name, qnty, price in products_data:
            product = self._find_or_create_product(name, created_products_names)
            self._create_sale_order_line(sale_order,product, qnty, price)
        
        # робимо нотатки
        notes_text = self._make_notes(skipped_missing_data_rows, invalid_data_rows, created_products_names)
        if notes_text:
            sale_order.write({'note': notes_text})

        return {
            "name":_("Order"),
            "type":"ir.actions.act_window",
            "res_model":"sale.order",
            "view_mode":"form",
            "res_id":sale_order.id,
            "target":"current",
        }
